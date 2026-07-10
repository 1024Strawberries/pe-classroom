from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db

router = APIRouter(prefix="/api/classes/{class_id}/students", tags=["students"])

HEADERS = {"学号": "student_no", "姓名": "name", "班级": "class_name"}


@router.get("", response_model=list[schemas.StudentOut])
def list_students(class_id: int, db: Session = Depends(get_db)):
    return (
        db.query(models.Student)
        .filter_by(course_class_id=class_id)
        .order_by(models.Student.position, models.Student.id)
        .all()
    )


@router.post("", response_model=schemas.StudentOut)
def create_student(class_id: int, data: schemas.StudentIn, db: Session = Depends(get_db)):
    if not db.get(models.CourseClass, class_id):
        raise HTTPException(status_code=404, detail="课程班级不存在")
    if not data.student_no.strip() or not data.name.strip():
        raise HTTPException(status_code=400, detail="学号和姓名不能为空")
    exists = db.query(models.Student).filter_by(course_class_id=class_id, student_no=data.student_no.strip()).first()
    if exists:
        raise HTTPException(status_code=400, detail="该学号已存在")
    max_position = (
        db.query(models.Student.position)
        .filter_by(course_class_id=class_id)
        .order_by(models.Student.position.desc())
        .first()
    )
    student = models.Student(
        course_class_id=class_id,
        student_no=data.student_no.strip(),
        name=data.name.strip(),
        class_name=data.class_name.strip(),
        position=(max_position[0] if max_position else 0) + 1,
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


@router.put("/{student_id}", response_model=schemas.StudentOut)
def update_student(class_id: int, student_id: int, data: schemas.StudentIn, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student or student.course_class_id != class_id:
        raise HTTPException(status_code=404, detail="学生不存在")
    duplicate = (
        db.query(models.Student)
        .filter_by(course_class_id=class_id, student_no=data.student_no.strip())
        .filter(models.Student.id != student_id)
        .first()
    )
    if duplicate:
        raise HTTPException(status_code=400, detail="该学号已存在")
    student.student_no = data.student_no.strip()
    student.name = data.name.strip()
    student.class_name = data.class_name.strip()
    db.commit()
    db.refresh(student)
    return student


@router.delete("/{student_id}")
def delete_student(class_id: int, student_id: int, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student or student.course_class_id != class_id:
        raise HTTPException(status_code=404, detail="学生不存在")
    db.delete(student)
    db.commit()
    return {"ok": True}


@router.post("/import")
async def import_students(class_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not db.get(models.CourseClass, class_id):
        raise HTTPException(status_code=404, detail="课程班级不存在")

    workbook = load_workbook(BytesIO(await file.read()))
    sheet = workbook.active
    header_row = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    indexes = {HEADERS[name]: idx for idx, name in enumerate(header_row) if name in HEADERS}
    if "student_no" not in indexes or "name" not in indexes:
        raise HTTPException(status_code=400, detail="Excel 必须包含学号、姓名列，可选班级列")

    imported = 0
    for position, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        student_no = str(row[indexes["student_no"]] or "").strip()
        name = str(row[indexes["name"]] or "").strip()
        if not student_no or not name:
            continue
        class_name = ""
        if "class_name" in indexes:
            class_name = str(row[indexes["class_name"]] or "").strip()
        student = db.query(models.Student).filter_by(course_class_id=class_id, student_no=student_no).first()
        if not student:
            student = models.Student(course_class_id=class_id, student_no=student_no)
            db.add(student)
        student.name = name
        student.class_name = class_name
        student.position = position
        imported += 1
    db.commit()
    return {"imported": imported}
